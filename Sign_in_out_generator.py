import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

print("Welcome to the Celsius and Beyond Sign-in/out System.  Created by Jared Sinasohn\n"
      "Before using this program, make sure this .exe file and the .csv file containing the week's campers are in the same folder"
      " on the computer.  Otherwise, the program will not work.\n")
csv_file = input("Enter the name of the csv file that contains the camper data below.  It should look like \"FILE NAME HERE.csv\" (NO QUOTATIONS)\n")
desired_name = input("Enter the desired name of the generated check-in/out sheet below.  It should look like \"DESIRED FILE NAME.xlsx\" (NO QUOTATIONS)\n")
try:
    df = pd.read_csv(csv_file)
except UnicodeDecodeError:
    df = pd.read_csv(csv_file, encoding="ISO-8859-1")
# Adjust column names as necessary
df.columns = [col.strip() for col in df.columns]  # Remove leading/trailing spaces

# Sort data by Session Name, then by First Name

try:
    df = df.sort_values(by=['Session Name', 'First Name'])
except KeyError:
    print("Please make sure all headers of the columns in the csv have no additional weird characters.  The headers should be as follows (case sensitive):\n"
          "First Name | Last Name | Extended Care | Session Name | Location")
    input("Press ENTER to end the program.  Once you have fixed the column names, rerun the .exe file")
    raise KeyError
# Create a new Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Summer Camp Roster"

# Define the dropdown options and colors
dropdown_options = ["Absent", "Dropped off", "Called down", "Sent down", "Picked up"]
colors = {
    "Absent": "ADD8E6",  # Baby blue
    "Dropped off": "D3D3D3",  # Light grey
    "Called down": "FF0000",  # Red
    "Sent down": "FFFF00",  # Yellow
    "Picked up": "008000"  # Green
}

# Create data validation for dropdown
dv = DataValidation(
    type="list",
    formula1='"Absent,Dropped off,Called down,Sent down,Picked up"',
    allow_blank=False
)
dv.error = 'Your entry is not in the list'
dv.errorTitle = 'Invalid Entry'
dv.prompt = 'Please select from the list'
dv.promptTitle = 'Dropdown List'

# Define borders
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# Iterate over the sorted data and add to the workbook
start_row = 1
for session_name, group in df.groupby('Session Name'):
    # Create table title with bold font and merge cells
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=5)
    title_cell = ws.cell(row=start_row, column=1, value=session_name)
    title_cell.font = Font(bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    start_row += 1

    # Add column headers with underlined font
    headers = ["First Name", "Last Name", "Status", "Time In", "Time Out"]
    for col_num, header in enumerate(headers, 1):
        header_cell = ws.cell(row=start_row, column=col_num, value=header)
        header_cell.font = Font(underline="single")
        header_cell.alignment = Alignment(horizontal='center', vertical='center')

    start_row += 1

    # Track the starting and ending rows of the current table
    table_start_row = start_row
    table_end_row = start_row + len(group) - 1

    # Add data rows
    for _, row in group.iterrows():
        first_name_cell = ws.cell(row=start_row, column=1, value=row["First Name"])
        last_name_cell = ws.cell(row=start_row, column=2, value=row["Last Name"])

        # Check if the kid is in extended care and italicize if text contains "PM"
        if pd.notna(row["Extended Care"]) and "pm" in row["Extended Care"].lower():
            first_name_cell.font = Font(italic=True)
            last_name_cell.font = Font(italic=True)

        ws.add_data_validation(dv)
        status_cell = ws.cell(row=start_row, column=3, value="Absent")
        dv.add(status_cell)
        status_cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=start_row, column=4, value="")
        ws.cell(row=start_row, column=5, value="")

        # Add thin borders between columns
        for col in range(1, 6):
            ws.cell(row=start_row, column=col).border = thin_border

        start_row += 1

    # Add a thick border around the entire table
    for row in range(table_start_row-2, start_row):
        ws.cell(row=row, column=1).border = Border(left=Side(style='thick'),right=Side(style='thin'))
        ws.cell(row=row, column=5).border = Border(right=Side(style='thick'),left=Side(style='thin'))

    for col in range(1, 6):
        if col == 1:
            ws.cell(row=table_start_row-2, column=col).border = Border(top=Side(style='thick'),left=Side(style='thick'),right=Side(style='thin'),bottom=Side(style='thin'))
            ws.cell(row=start_row-1, column=col).border = Border(bottom=Side(style='thick'),left=Side(style='thick'),right=Side(style='thin'))
        elif col == 5:
            ws.cell(row=table_start_row - 2, column=col).border = Border(top=Side(style='thick'),right=Side(style='thick'),left=Side(style='thin'),bottom=Side(style='thin'))
            ws.cell(row=start_row - 1, column=col).border = Border(bottom=Side(style='thick'),right=Side(style='thick'),left=Side(style='thin'))
        else:
            ws.cell(row=table_start_row - 2, column=col).border = Border(top=Side(style='thick'),bottom=Side(style='thin'))
            ws.cell(row=start_row - 1, column=col).border = Border(bottom=Side(style='thick'),right=Side(style='thin'))

    # Add an extra row for separation
    start_row += 1

# Apply the coloring for dropdown selections
for opt, color in colors.items():
    formula = f'SEARCH("{opt}",C2:C{start_row})'
    dxf = DifferentialStyle(fill=PatternFill(start_color=color, end_color=color, fill_type="solid"))
    rule = Rule(type="expression", dxf=dxf, formula=[formula])
    ws.conditional_formatting.add(f'C2:C{start_row}', rule)

# Save the workbook
wb.save(desired_name)
print(f"Excel file saved as {desired_name}")